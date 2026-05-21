"""
对账系统 — Excel 导出 Service
=========================================
职责（Requirements 7.1, 7.3, 7.5, 7.6）：
  1. 生成单份对账单 Excel（.xlsx）：包含对账单编号、供应商名称、周期、
     行项明细及汇总金额（Requirement 7.1）
  2. 批量导出多份对账单（每份占一个 sheet）（Requirement 7.3）
  3. 导出异常报告：包含全部 Anomaly 的详细信息及处理状态（Requirement 7.5）
  4. 导出失败时通知操作人员并记录失败原因（Requirement 7.6）

设计要点：
  - 使用 openpyxl 直接构造 Workbook，便于控制表头样式与列宽
  - 所有方法返回 ``ExportResult`` 数据类，包含文件名 / mimetype / 二进制内容，
    Controller 可据此返回 StreamingResponse
  - 失败时抛出 ServiceException，并通过 NotificationService 通知操作人，
    同时通过 loguru.logger.exception 留下完整堆栈
  - 与 ReconciliationAuditService 集成（log_action_safe）：
    单份/批量/异常报告导出成功后写入 entity_type='statement', action='export'
    的审计日志，失败不影响主流程

通知与失败处理（Requirement 7.6）：
  - 任何步骤抛错先 ``logger.exception`` 记录原因
  - 调用 ReconciliationNotificationService 的渠道直接发送 ``[导出失败]`` 通知
    给操作人员（recipient = ``user:{operator_id}``，由渠道按账号 fan-out）
  - 然后再抛出 ServiceException 让上层捕获
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Iterable, Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from exceptions.exception import ServiceException
from module_entrust.entity.do.entrust_do import EntrustSupplier
from module_entrust.entity.do.reconciliation_do import (
    Anomaly,
    ReconciliationLineItem,
    ReconciliationStatement,
)
from module_entrust.service.reconciliation_audit_service import (
    ACTION_EXPORT,
    ENTITY_TYPE_STATEMENT,
    ReconciliationAuditService,
)
from module_entrust.service.reconciliation_notification_service import (
    ReconciliationNotificationService,
)


# ---------------------------------------------------------------------------
# 常量与样式
# ---------------------------------------------------------------------------

XLSX_MIME_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

# 表头样式
_HEADER_FILL = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 摘要区样式
_SUMMARY_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
_SUMMARY_FONT = Font(bold=True)

# 单元格通用对齐
_CELL_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
_CELL_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# Excel sheet 名称的非法字符 — 用下划线代替
_SHEET_NAME_INVALID = set(r':\/?*[]')
_SHEET_NAME_MAX_LEN = 31  # openpyxl 限制

# 行项表头（顺序对应 Requirement 1.2 中的字段）
_LINE_ITEM_HEADERS: tuple[str, ...] = (
    '序号',
    '委外单号',
    '工序名称',
    '零件编号',
    '零件名称',
    '单价',
    '数量',
    '行项金额',
)

# 异常报告表头（覆盖 Requirement 3.x 关键字段）
_ANOMALY_HEADERS: tuple[str, ...] = (
    '序号',
    '异常ID',
    '对账单编号',
    '行项ID',
    '异常类型',
    '严重程度',
    '差异金额',
    '处理状态',
    '异常描述',
    '解决人ID',
    '解决时间',
    '创建时间',
    '更新时间',
)


# ---------------------------------------------------------------------------
# 数据载体
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ExportResult:
    """导出结果 — Controller 据此返回 StreamingResponse。"""
    file_name: str
    mime_type: str
    content: bytes

    @property
    def file_size(self) -> int:
        return len(self.content)


# ---------------------------------------------------------------------------
# 内部工具
# ---------------------------------------------------------------------------

def _safe_sheet_name(raw: str, fallback: str) -> str:
    """生成合法的 sheet 名称 (≤31 字符，去除非法字符)。"""
    if not raw:
        raw = fallback
    cleaned = ''.join('_' if c in _SHEET_NAME_INVALID else c for c in str(raw))
    cleaned = cleaned.strip() or fallback
    return cleaned[:_SHEET_NAME_MAX_LEN]


def _decimal_or_zero(value) -> Decimal:
    """把可能为 None / Decimal / 数字 / 字符串的金额规范为 Decimal（None → 0）。"""
    if value is None:
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:  # noqa: BLE001
        return Decimal('0')


def _fmt_amount(value) -> str:
    """金额格式化为带两位小数的字符串。"""
    return f'{_decimal_or_zero(value):.2f}'


def _fmt_dt(value: Optional[datetime]) -> str:
    """日期时间格式化为 YYYY-MM-DD HH:MM:SS。"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return str(value)


def _apply_header_row(ws: Worksheet, headers: Iterable[str], row: int = 1) -> None:
    """写入并美化表头行。"""
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """根据当前数据估算列宽（简单实现：取每列最长字符数 + 2）。"""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class ReconciliationExportService:
    """对账单 Excel 导出服务。"""

    # ------------------------------------------------------------------
    # 数据加载
    # ------------------------------------------------------------------

    @staticmethod
    async def _load_statement(
        db: AsyncSession, statement_id: int
    ) -> ReconciliationStatement:
        stmt = select(ReconciliationStatement).where(
            ReconciliationStatement.id == statement_id
        )
        statement = (await db.execute(stmt)).scalar_one_or_none()
        if statement is None:
            raise ServiceException(message=f'对账单不存在: id={statement_id}')
        return statement

    @staticmethod
    async def _load_line_items(
        db: AsyncSession, statement_id: int
    ) -> list[ReconciliationLineItem]:
        stmt = (
            select(ReconciliationLineItem)
            .where(ReconciliationLineItem.statement_id == statement_id)
            .order_by(ReconciliationLineItem.id.asc())
        )
        return list((await db.execute(stmt)).scalars().all())

    @staticmethod
    async def _load_supplier_name(
        db: AsyncSession, supplier_id: int
    ) -> str:
        if not supplier_id:
            return ''
        name = await db.scalar(
            select(EntrustSupplier.name).where(EntrustSupplier.id == supplier_id)
        )
        return name or f'供应商#{supplier_id}'

    # ------------------------------------------------------------------
    # 单份对账单 Sheet 写入
    # ------------------------------------------------------------------

    @staticmethod
    def _write_statement_sheet(
        ws: Worksheet,
        statement: ReconciliationStatement,
        supplier_name: str,
        line_items: list[ReconciliationLineItem],
    ) -> None:
        """
        在给定 sheet 内写入一份对账单（meta 区 + 行项明细 + 汇总）。

        布局：
          A1:B1  对账单编号
          A2:B2  供应商
          A3:B3  对账周期
          A4:B4  状态 / 确认状态
          A5:B5  导出时间
          (空行)
          row=7  行项表头
          row=8+ 行项数据
          末尾   汇总行（合计 = Σ 行项金额）
        """
        # ---------------------- meta 区 ----------------------
        meta_rows: list[tuple[str, str]] = [
            ('对账单编号', statement.statement_no or ''),
            ('供应商', f'{supplier_name} (ID={statement.supplier_id})'),
            (
                '对账周期',
                f'{statement.period_start} 至 {statement.period_end}',
            ),
            (
                '状态 / 确认状态',
                f'{statement.status} / {statement.confirmation_status}',
            ),
            ('导出时间', _fmt_dt(datetime.now())),
        ]
        for r, (label, value) in enumerate(meta_rows, start=1):
            label_cell = ws.cell(row=r, column=1, value=label)
            label_cell.font = _SUMMARY_FONT
            label_cell.fill = _SUMMARY_FILL
            label_cell.alignment = _CELL_ALIGN
            value_cell = ws.cell(row=r, column=2, value=value)
            value_cell.alignment = _CELL_ALIGN
            # 让 meta value 跨到行项末尾，提升可读性
            ws.merge_cells(
                start_row=r,
                start_column=2,
                end_row=r,
                end_column=len(_LINE_ITEM_HEADERS),
            )

        # ---------------------- 行项表头 ----------------------
        header_row = len(meta_rows) + 2  # 留一空行
        _apply_header_row(ws, _LINE_ITEM_HEADERS, row=header_row)

        # ---------------------- 行项数据 ----------------------
        line_total: Decimal = Decimal('0')
        cur_row = header_row + 1
        for idx, li in enumerate(line_items, start=1):
            amount = _decimal_or_zero(li.total_amount)
            line_total += amount
            ws.cell(row=cur_row, column=1, value=idx).alignment = _CELL_ALIGN_RIGHT
            ws.cell(row=cur_row, column=2, value=li.order_no or '').alignment = _CELL_ALIGN
            ws.cell(row=cur_row, column=3, value=li.process_name or '').alignment = _CELL_ALIGN
            ws.cell(row=cur_row, column=4, value=li.part_no or '').alignment = _CELL_ALIGN
            ws.cell(row=cur_row, column=5, value=li.part_name or '').alignment = _CELL_ALIGN
            unit_price_cell = ws.cell(
                row=cur_row, column=6,
                value=float(_decimal_or_zero(li.unit_price)) if li.unit_price is not None else None,
            )
            unit_price_cell.number_format = '#,##0.00'
            unit_price_cell.alignment = _CELL_ALIGN_RIGHT
            qty_cell = ws.cell(row=cur_row, column=7, value=li.quantity)
            qty_cell.alignment = _CELL_ALIGN_RIGHT
            amount_cell = ws.cell(row=cur_row, column=8, value=float(amount))
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = _CELL_ALIGN_RIGHT
            cur_row += 1

        # ---------------------- 汇总行 ----------------------
        summary_row = cur_row
        # 行项汇总（Σ 行项 total_amount）
        label_cell = ws.cell(row=summary_row, column=1, value='行项合计')
        label_cell.font = _SUMMARY_FONT
        label_cell.fill = _SUMMARY_FILL
        label_cell.alignment = _CELL_ALIGN
        ws.merge_cells(
            start_row=summary_row, start_column=1,
            end_row=summary_row, end_column=7,
        )
        line_sum_cell = ws.cell(row=summary_row, column=8, value=float(line_total))
        line_sum_cell.font = _SUMMARY_FONT
        line_sum_cell.fill = _SUMMARY_FILL
        line_sum_cell.number_format = '#,##0.00'
        line_sum_cell.alignment = _CELL_ALIGN_RIGHT

        # 对账单上记录的 total_amount（与行项 sum 一般一致；
        # Property 2 保证）— 同时输出便于核对
        stmt_total_row = summary_row + 1
        label_cell2 = ws.cell(row=stmt_total_row, column=1, value='对账单汇总金额')
        label_cell2.font = _SUMMARY_FONT
        label_cell2.fill = _SUMMARY_FILL
        label_cell2.alignment = _CELL_ALIGN
        ws.merge_cells(
            start_row=stmt_total_row, start_column=1,
            end_row=stmt_total_row, end_column=7,
        )
        stmt_total_cell = ws.cell(
            row=stmt_total_row, column=8,
            value=float(_decimal_or_zero(statement.total_amount)),
        )
        stmt_total_cell.font = _SUMMARY_FONT
        stmt_total_cell.fill = _SUMMARY_FILL
        stmt_total_cell.number_format = '#,##0.00'
        stmt_total_cell.alignment = _CELL_ALIGN_RIGHT

        # ---------------------- 列宽自适应 ----------------------
        _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 1. 单份导出（Requirement 7.1）
    # ------------------------------------------------------------------

    @staticmethod
    async def export_statement_excel(
        db: AsyncSession,
        statement_id: int,
        operator_id: int = 0,
    ) -> ExportResult:
        """
        生成单份对账单 .xlsx 文件。

        Args:
            db: AsyncSession
            statement_id: 对账单 ID
            operator_id: 操作人 ID（用于失败通知；0 表示系统）

        Returns:
            ExportResult(file_name, mime_type, content)

        Raises:
            ServiceException: 对账单不存在 / 生成失败
        """
        try:
            statement = await ReconciliationExportService._load_statement(
                db, statement_id
            )
            line_items = await ReconciliationExportService._load_line_items(
                db, statement_id
            )
            supplier_name = await ReconciliationExportService._load_supplier_name(
                db, statement.supplier_id
            )

            wb = Workbook()
            ws = wb.active
            ws.title = _safe_sheet_name(
                statement.statement_no or f'对账单_{statement_id}',
                fallback=f'对账单_{statement_id}',
            )
            ReconciliationExportService._write_statement_sheet(
                ws, statement, supplier_name, line_items
            )

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.getvalue()

            file_name = (
                f'{statement.statement_no or f"reconciliation_{statement_id}"}.xlsx'
            )
            logger.info(
                '[ReconciliationExportService] 导出对账单 Excel 成功 '
                'statement_id={} statement_no={} size={}B operator={}',
                statement.id, statement.statement_no, len(content), operator_id,
            )
            # 审计日志：对账单导出（Requirement 8.1）— 失败不影响主流程
            await ReconciliationAuditService.log_action_safe(
                db=db,
                entity_type=ENTITY_TYPE_STATEMENT,
                entity_id=statement.id,
                action=ACTION_EXPORT,
                operator_id=int(operator_id or 0),
                detail={
                    'sub_action': 'export_statement_excel',
                    'statement_no': statement.statement_no,
                    'file_name': file_name,
                    'file_size': len(content),
                },
                autocommit=True,
            )
            return ExportResult(
                file_name=file_name,
                mime_type=XLSX_MIME_TYPE,
                content=content,
            )
        except ServiceException:
            # 数据缺失等显式错误：直接通知 + 抛出
            await ReconciliationExportService._notify_export_failure(
                db,
                operator_id=operator_id,
                kind='statement_excel',
                target=str(statement_id),
                reason='对账单不存在或参数非法',
            )
            raise
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                '[ReconciliationExportService] 导出对账单 Excel 失败 '
                'statement_id={} operator={}',
                statement_id, operator_id,
            )
            await ReconciliationExportService._notify_export_failure(
                db,
                operator_id=operator_id,
                kind='statement_excel',
                target=str(statement_id),
                reason=str(exc),
            )
            raise ServiceException(
                message=f'对账单导出失败: statement_id={statement_id}, 原因={exc}'
            )

    # ------------------------------------------------------------------
    # 2. 批量导出（Requirement 7.3）
    # ------------------------------------------------------------------

    @staticmethod
    async def export_batch_excel(
        db: AsyncSession,
        statement_ids: list[int],
        operator_id: int = 0,
    ) -> ExportResult:
        """
        批量导出多份对账单到同一份 .xlsx（每份占一个 sheet）。

        Args:
            db: AsyncSession
            statement_ids: 对账单 ID 列表（去重后按输入顺序处理）
            operator_id: 操作人 ID

        Returns:
            ExportResult — 默认文件名 ``reconciliation_batch_{YYYYMMDDHHMMSS}.xlsx``

        Raises:
            ServiceException: 输入为空 / 任一对账单不存在 / 生成失败
        """
        if not statement_ids:
            raise ServiceException(message='批量导出的对账单ID列表不能为空')

        # 去重保留顺序
        seen: set[int] = set()
        ordered_ids: list[int] = []
        for sid in statement_ids:
            if sid in seen:
                continue
            seen.add(sid)
            ordered_ids.append(sid)

        try:
            wb = Workbook()
            # 移除默认 sheet，等下逐份添加
            default = wb.active
            wb.remove(default)

            # sheet 名重名处理
            used_names: set[str] = set()
            for sid in ordered_ids:
                statement = await ReconciliationExportService._load_statement(
                    db, sid
                )
                line_items = await ReconciliationExportService._load_line_items(
                    db, sid
                )
                supplier_name = await ReconciliationExportService._load_supplier_name(
                    db, statement.supplier_id
                )

                base_name = _safe_sheet_name(
                    statement.statement_no or f'对账单_{sid}',
                    fallback=f'对账单_{sid}',
                )
                # 确保名字唯一
                name = base_name
                suffix = 1
                while name in used_names:
                    suffix += 1
                    candidate = f'{base_name}_{suffix}'
                    name = candidate[-_SHEET_NAME_MAX_LEN:]
                used_names.add(name)

                ws = wb.create_sheet(title=name)
                ReconciliationExportService._write_statement_sheet(
                    ws, statement, supplier_name, line_items
                )

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.getvalue()

            ts = datetime.now().strftime('%Y%m%d%H%M%S')
            file_name = f'reconciliation_batch_{ts}.xlsx'

            logger.info(
                '[ReconciliationExportService] 批量导出对账单 Excel 成功 '
                'count={} size={}B operator={}',
                len(ordered_ids), len(content), operator_id,
            )
            # 审计日志：批量导出（每份对账单各写一条）— 失败不影响主流程
            for sid in ordered_ids:
                await ReconciliationAuditService.log_action_safe(
                    db=db,
                    entity_type=ENTITY_TYPE_STATEMENT,
                    entity_id=sid,
                    action=ACTION_EXPORT,
                    operator_id=int(operator_id or 0),
                    detail={
                        'sub_action': 'export_batch_excel',
                        'batch_size': len(ordered_ids),
                        'file_name': file_name,
                    },
                    autocommit=True,
                )
            return ExportResult(
                file_name=file_name,
                mime_type=XLSX_MIME_TYPE,
                content=content,
            )
        except ServiceException:
            await ReconciliationExportService._notify_export_failure(
                db,
                operator_id=operator_id,
                kind='statement_batch_excel',
                target=','.join(str(s) for s in ordered_ids),
                reason='对账单不存在或参数非法',
            )
            raise
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                '[ReconciliationExportService] 批量导出对账单 Excel 失败 '
                'ids={} operator={}',
                ordered_ids, operator_id,
            )
            await ReconciliationExportService._notify_export_failure(
                db,
                operator_id=operator_id,
                kind='statement_batch_excel',
                target=','.join(str(s) for s in ordered_ids),
                reason=str(exc),
            )
            raise ServiceException(message=f'批量对账单导出失败: 原因={exc}')

    # ------------------------------------------------------------------
    # 3. 异常报告导出（Requirement 7.5）
    # ------------------------------------------------------------------

    @staticmethod
    async def export_anomaly_report(
        db: AsyncSession,
        statement_id: int,
        operator_id: int = 0,
    ) -> ExportResult:
        """
        生成单份对账单的异常报告 .xlsx。

        - 列出该对账单下全部 Anomaly 的详细信息及处理状态（包括 critical/warning/info）
        - 仅一个 sheet，按 anomaly.id 升序

        Args:
            db: AsyncSession
            statement_id: 对账单 ID
            operator_id: 操作人 ID

        Returns:
            ExportResult — 文件名 ``anomaly_report_{statement_no}.xlsx``

        Raises:
            ServiceException: 对账单不存在 / 生成失败
        """
        try:
            statement = await ReconciliationExportService._load_statement(
                db, statement_id
            )
            supplier_name = await ReconciliationExportService._load_supplier_name(
                db, statement.supplier_id
            )

            anomaly_stmt = (
                select(Anomaly)
                .where(Anomaly.statement_id == statement_id)
                .order_by(Anomaly.id.asc())
            )
            anomalies: list[Anomaly] = list(
                (await db.execute(anomaly_stmt)).scalars().all()
            )

            wb = Workbook()
            ws = wb.active
            ws.title = _safe_sheet_name(
                f'异常_{statement.statement_no or statement_id}',
                fallback=f'异常_{statement_id}',
            )

            # 头部摘要
            meta_rows = [
                ('对账单编号', statement.statement_no or ''),
                ('供应商', f'{supplier_name} (ID={statement.supplier_id})'),
                (
                    '对账周期',
                    f'{statement.period_start} 至 {statement.period_end}',
                ),
                ('异常总数', str(len(anomalies))),
                ('导出时间', _fmt_dt(datetime.now())),
            ]
            for r, (label, value) in enumerate(meta_rows, start=1):
                lc = ws.cell(row=r, column=1, value=label)
                lc.font = _SUMMARY_FONT
                lc.fill = _SUMMARY_FILL
                lc.alignment = _CELL_ALIGN
                vc = ws.cell(row=r, column=2, value=value)
                vc.alignment = _CELL_ALIGN
                ws.merge_cells(
                    start_row=r, start_column=2,
                    end_row=r, end_column=len(_ANOMALY_HEADERS),
                )

            # 表头
            header_row = len(meta_rows) + 2
            _apply_header_row(ws, _ANOMALY_HEADERS, row=header_row)

            # 数据行
            cur_row = header_row + 1
            for idx, a in enumerate(anomalies, start=1):
                ws.cell(row=cur_row, column=1, value=idx).alignment = _CELL_ALIGN_RIGHT
                ws.cell(row=cur_row, column=2, value=a.id).alignment = _CELL_ALIGN_RIGHT
                ws.cell(row=cur_row, column=3, value=statement.statement_no or '').alignment = _CELL_ALIGN
                ws.cell(row=cur_row, column=4, value=a.line_item_id).alignment = _CELL_ALIGN_RIGHT
                ws.cell(row=cur_row, column=5, value=a.anomaly_type or '').alignment = _CELL_ALIGN
                ws.cell(row=cur_row, column=6, value=a.severity or '').alignment = _CELL_ALIGN
                diff_cell = ws.cell(
                    row=cur_row, column=7,
                    value=(
                        float(_decimal_or_zero(a.diff_amount))
                        if a.diff_amount is not None else None
                    ),
                )
                diff_cell.number_format = '#,##0.00'
                diff_cell.alignment = _CELL_ALIGN_RIGHT
                ws.cell(row=cur_row, column=8, value=a.status or '').alignment = _CELL_ALIGN
                ws.cell(row=cur_row, column=9, value=a.description or '').alignment = _CELL_ALIGN
                ws.cell(row=cur_row, column=10, value=a.resolved_by).alignment = _CELL_ALIGN_RIGHT
                ws.cell(row=cur_row, column=11, value=_fmt_dt(a.resolved_at)).alignment = _CELL_ALIGN
                ws.cell(row=cur_row, column=12, value=_fmt_dt(a.created_at)).alignment = _CELL_ALIGN
                ws.cell(row=cur_row, column=13, value=_fmt_dt(a.updated_at)).alignment = _CELL_ALIGN
                cur_row += 1

            _autosize_columns(ws)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.getvalue()

            file_name = (
                f'anomaly_report_'
                f'{statement.statement_no or f"reconciliation_{statement_id}"}.xlsx'
            )
            logger.info(
                '[ReconciliationExportService] 导出异常报告成功 '
                'statement_id={} count={} size={}B operator={}',
                statement.id, len(anomalies), len(content), operator_id,
            )
            # 审计日志：异常报告导出（Requirement 8.1）— 失败不影响主流程
            await ReconciliationAuditService.log_action_safe(
                db=db,
                entity_type=ENTITY_TYPE_STATEMENT,
                entity_id=statement.id,
                action=ACTION_EXPORT,
                operator_id=int(operator_id or 0),
                detail={
                    'sub_action': 'export_anomaly_report',
                    'statement_no': statement.statement_no,
                    'anomaly_count': len(anomalies),
                    'file_name': file_name,
                },
                autocommit=True,
            )
            return ExportResult(
                file_name=file_name,
                mime_type=XLSX_MIME_TYPE,
                content=content,
            )
        except ServiceException:
            await ReconciliationExportService._notify_export_failure(
                db,
                operator_id=operator_id,
                kind='anomaly_report',
                target=str(statement_id),
                reason='对账单不存在或参数非法',
            )
            raise
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                '[ReconciliationExportService] 导出异常报告失败 '
                'statement_id={} operator={}',
                statement_id, operator_id,
            )
            await ReconciliationExportService._notify_export_failure(
                db,
                operator_id=operator_id,
                kind='anomaly_report',
                target=str(statement_id),
                reason=str(exc),
            )
            raise ServiceException(
                message=f'异常报告导出失败: statement_id={statement_id}, 原因={exc}'
            )

    # ------------------------------------------------------------------
    # 失败通知（Requirement 7.6）
    # ------------------------------------------------------------------

    @staticmethod
    async def _notify_export_failure(
        db: AsyncSession,
        operator_id: int,
        kind: str,
        target: str,
        reason: str,
    ) -> None:
        """
        通过通知渠道告知操作人员导出失败，并记录失败原因。

        - 不再抛出异常，避免遮蔽原始失败原因
        - 操作人未知（0）时尝试通知财务（与异常通知一致）

        Args:
            kind: 失败的导出类型，如 ``statement_excel`` /
                  ``statement_batch_excel`` / ``anomaly_report``
            target: 目标实体 ID（或逗号分隔的批量 ID 列表）
            reason: 失败原因（异常 message）
        """
        try:
            channel = ReconciliationNotificationService.get_channel()
            if operator_id and operator_id > 0:
                recipient = f'user:{operator_id}'
            else:
                recipient = 'role:finance'

            subject = f'[导出失败] {kind} 目标={target}'
            body = (
                f'对账系统导出任务失败：\n'
                f'  类型: {kind}\n'
                f'  目标: {target}\n'
                f'  操作人ID: {operator_id or "system"}\n'
                f'  失败时间: {_fmt_dt(datetime.now())}\n'
                f'  失败原因: {reason}'
            )
            metadata = {
                'kind': kind,
                'target': target,
                'operator_id': operator_id,
                'reason': reason,
            }
            await channel.send(recipient, subject, body, metadata)
        except Exception as exc:  # noqa: BLE001
            # 二次失败：仅记录，不再向上抛出
            logger.error(
                '[ReconciliationExportService] 导出失败通知发送失败 '
                'kind={} target={} reason={} notify_err={}',
                kind, target, reason, exc,
            )


__all__ = [
    'ExportResult',
    'ReconciliationExportService',
    'XLSX_MIME_TYPE',
]
